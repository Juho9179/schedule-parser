################
#
#   Schedule-parser v 0.1
#
#   python3 .\schedule-parser.py NAME '.\schedule.xlsx'
#   -> creates a new file, in the same directory as .xlsx file with same name, but .csv extension.
#   -> OVERWRITES old file, should it exist.
#
################
import openpyxl
import sys
year_variable = "20"
date_delimiter = "/"
filename = ""
employee_name = ""
workbook = ""
ws = ""
headers = ["Subject", '"Start Date"', '"End Date"', "Private"]
private = "true"

# saves csv data to file
def save_csv(csv, output_filename):
    with open(output_filename, "w") as output:
        print(csv, file=output)

# gets CSV from data list of shifts
def get_csv(headers, data):
    csv = ''
    # import headers
    for i in headers:
        csv += i + ","
    # remove tailing ,
    csv = csv[:-1] + "\n"
    # add shifts
    for shift in data:
        csv += '"' + shift["detail"] + '",' + shift["date"] +',' + shift["date"] + "," + private + "\n"
    return csv

# Initialises file name and employee name.
def init():
    global filename
    global employee_name
    global workbook
    global ws
    argument_list = sys.argv[1:]
    for i in argument_list:
        if i.endswith(".xlsx"):
            filename = argument_list.pop(argument_list.index(i))
    employee_name = ' '.join(argument_list)
    
    workbook = openpyxl.load_workbook(filename)
    ws = workbook.active

# Returns shift for specific employee on a specific date
# Returns
# {'date': '15/06/21', 'detail': '0:00-8:00 '}
def get_shift(employee_row_info, date_info):
    global ws
    # checks employee row info according to date
    # checks whether row_info is dual or singular row and returns information accordingly
    shift = {}
    shift["date"] = date_info[0]
    # check if no shift, if no shift: return None
    actual_shift = ws.cell(employee_row_info["row"], date_info[1]).internal_value
    if (actual_shift == None):
        return None
    shift["detail"] = ws.cell(employee_row_info["row"], date_info[1]).internal_value.strip()
    # check whether shift is dual, if dual: concatenate
    if ((employee_row_info["dual"] == True) and (ws.cell(employee_row_info["row"] + 1, date_info[1] != None))):
        # check whether there is data in row below
        shift["detail"] += " " + ws.cell(employee_row_info["row"] + 1, date_info[1]).internal_value.strip()

    if (shift["detail"] == " "):
        return None
    return shift

# Gets correct dates for each day in schedule
# Returns list
# ['14/06/21', '15/06/21', '16/06/21', '17/06/21', '18/06/21', '19/06/21', '20/06/21', '21/06/21', '22/06/21', '23/06/21', '24/06/21', '25/06/21', '26/06/21', '27/06/21', '28/06/21', '29/06/21', '30/06/21', '1/07/21', '2/07/21', '3/07/21', '4/07/21']
def get_dates():
    global ws
    shift_dates = []
    date_range = ws.cell(1, 16).internal_value
    dates = date_range.split("-")
    start_date = dates[0]
    end_date = dates[1]

    start_month = start_date.split(".")[1]
    end_month = end_date.split(".")[1]

    start_year = start_date.split(".")[2]
    end_year = end_date.split(".")[2]

    initial_date = ws.cell(3, 3).internal_value
    curr_date = 0
    counter = 3
    while curr_date != None:
        curr_date = ws.cell(3, counter).internal_value
        curr_date_column = ws.cell(3, counter).column
        if (curr_date == None):
            break
        if curr_date < initial_date:
            shift_dates.append([str(curr_date) + date_delimiter + str(end_month) + date_delimiter + str(end_year), curr_date_column])
        else:
            shift_dates.append([str(curr_date) + date_delimiter + str(start_month) + date_delimiter + str(start_year), curr_date_column])
        counter = counter + 1
    return shift_dates

# Returns row number for employee and checks whether dual row or single-row
# Returns object:
# {'coordinate': 'B15', 'column': 2, 'row': 15, 'dual': False}
# Returns False if employee is not found.
def get_row_info(employee_name):
    global ws
    row_info = {}
    row_info["coordinate"] = ""
    row_info["column"] = ""
    row_info["row"] = 0
    row_info["dual"] = False

    for row in ws.iter_rows():
        for cell in row:
            # check if cell contents match employee name
            if (cell.internal_value == employee_name):
                row_info["coordinate"] = cell.coordinate
                row_info["column"] = cell.column
                row_info["row"] = cell.row
                break
                
    # Check for errors
    try:
        # check whether row below is another employee or empty: singular or dual row
        next_entry = ws.cell(row_info["row"] + 1, row_info["column"]).internal_value
        if (next_entry == None):
            row_info["dual"] = True
        elif (next_entry.strip() == ""):
            row_info["dual"] = True

    except TypeError:
        print("ERROR: Employee " + employee_name + " not found.")
        return False

    return row_info

def get_all_shifts(dates_list, employee_name):
    all_shifts = []
    for i in dates_list:
        shift = get_shift(get_row_info(employee_name), i)
        if (shift != None):
            all_shifts.append(shift)
    return all_shifts

def main():
    init()
    save_csv(get_csv(headers, get_all_shifts(get_dates(), employee_name)), filename + ".csv")

if __name__ == "__main__":
    main()